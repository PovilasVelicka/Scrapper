from src.interfaces import repository as repo
from openpyxl import Workbook, load_workbook
from pathlib import Path
from typing import Optional

from src.interfaces.logger import ILogger


class ExcelRepository(repo.IDataAccessRepository):
    def __init__(self, file_path: str, logger: ILogger):
        self.__logger = logger
        self.__logger.log_debug(f"Repository init successfully")
        self.file_path = file_path
        self._ensure_workbook()


    def _ensure_workbook(self):
        if not Path(self.file_path).exists():
            self.__logger.log_debug(f"def:ensure_workbook - workbook not exists, trying to create new xlsx workbook")
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "products"
            ws1.append(["id", "name", "description", "price"])
            ws2 = wb.create_sheet("product_details")
            ws2.append(["product_id", "key", "value"])
            try:
                wb.save(self.file_path)
                self.__logger.log_debug(f"def:ensure_workbook - workbook created")
            except Exception as e:
                self.__logger.log_error(f"def:ensure_workbook - error: {e}")


    def get_first_or_default(self, id_keys: dict[str, str | int]) -> Optional[dict]:
        wb = load_workbook(self.file_path)
        ws_products = wb["products"]
        ws_details = wb["product_details"]

        for row in ws_products.iter_rows(min_row=2, values_only=True):
            if row[0] == id_keys["id"]:
                product = {
                    "id": row[0],
                    "name": row[1],
                    "description": row[2],
                    "price": row[3],
                    "details": []
                }
                for d_row in ws_details.iter_rows(min_row=2, values_only=True):
                    if d_row[0] == row[0]:
                        product["details"].append({d_row[1]: d_row[2]})
                return product
        return None


    def update(self, new_item: dict, id_keys: dict[str, str | int]) -> dict:
        wb = load_workbook(self.file_path)
        ws_products = wb["products"]
        ws_details = wb["product_details"]

        # Update main product
        for row in ws_products.iter_rows(min_row=2):
            if row[0].value == id_keys["id"]:
                row[1].value = new_item["name"]
                row[2].value = new_item["description"]
                row[3].value = new_item["price"]
                break

        # Delete old details
        details_to_keep = []
        for row in ws_details.iter_rows(min_row=2):
            if row[0].value != id_keys["id"]:
                details_to_keep.append([cell.value for cell in row])

        # Clear sheet and re-add header + retained rows
        ws_details.delete_rows(2, ws_details.max_row)
        for row in details_to_keep:
            ws_details.append(row)

        # Add new details
        for detail in new_item.get("details", []):
            for key, value in detail.items():
                ws_details.append([id_keys["id"], key, value])
        try:
            wb.save(self.file_path)
            self.__logger.log_debug(f"def:update - item {id_keys} update successfully")
        except Exception as e:
            self.__logger.log_error(f"def:update - error: {e}")
        return new_item


    def insert(self, item: dict) -> dict:
        wb = load_workbook(self.file_path)
        ws_products = wb["products"]
        ws_details = wb["product_details"]

        ws_products.append([
            item["id"],
            item["name"],
            item["description"],
            item["price"]
        ])

        for detail in item.get("details", []):
            for key, value in detail.items():
                ws_details.append([item["id"], key, value])

        try:
            wb.save(self.file_path)
            self.__logger.log_debug(f"def:insert - successfully")
        except Exception as e:
            self.__logger.log_error(f"def:insert - error: {e}")
        return item